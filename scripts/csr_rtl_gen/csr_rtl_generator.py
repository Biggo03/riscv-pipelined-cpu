#!/usr/bin/env python3
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
import yaml
import sys
import os

def excel_to_yaml(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    
    reg_data = {}

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(headers, row))
        name = entry.pop("CSR_NAME").upper()
        reg_data[name] = entry
        reg_data[name]["ADDR_MACRO"] = f"{name.upper()}_ADDR"

    with open(output_file, "w") as f:
        yaml_str = yaml.dump(reg_data, sort_keys=False)
        yaml_str = yaml_str.replace("\n- ", "\n\n- ")

        f.write(yaml_str)

    return

def generate_macros(yaml_file, macro_file):
    with open(yaml_file, "r") as f:
        reg_data = yaml.safe_load(f)

    with open(macro_file, "w") as f:
        f.write("////////////////////////////////////////\n")
        f.write("// AUTO-GENERATED CSR REGISTER MACROS //\n")
        f.write("////////////////////////////////////////\n")
        
        # figure out longest CSR name for alignment
        max_name_len = max(len(csr_name) for csr_name in reg_data.keys())

        for csr_info in reg_data.values():
            name_field = csr_info['ADDR_MACRO'].ljust(max_name_len + 10)
            addr_field = f"12'h{csr_info['ADDRESS']}"
            line = f"`define {name_field} {addr_field} // {csr_info['DESCRIPTION']}\n"
            f.write(line)

    return

def generate_rtl(yaml_file, rtl_op_path, tb_op_path):
    with open(yaml_file, "r") as f:
        reg_data = yaml.safe_load(f)

    for csr_name, csr_info in reg_data.items():
        reg_data[csr_name]["RESET_VAL"] = f"32'h{csr_info['RESET_VAL']}"
        reg_data[csr_name]["ADDR_MACRO"] = f"`{csr_info['ADDR_MACRO']}"


    max_macro_len = max(len(csr_info["ADDR_MACRO"]) for csr_info in reg_data.values())
    max_name_len = max(len(name) for name in reg_data.keys()) 
    
    # Jinja setup
    env = Environment(loader=FileSystemLoader("templates"))

    rtl_template = env.get_template("csr_regfile_template.sv.j2")
    rtl_output = rtl_template.render(csr_data=reg_data, max_name_len=max_name_len, max_macro_len=max_macro_len)

    tb_template = env.get_template("csr_regfile_tb_template.sv.j2")
    tb_output = tb_template.render(csr_data=reg_data, max_name_len=max_name_len, max_macro_len=max_macro_len)

    with open(rtl_op_path, "w") as rtl_file, open(tb_op_path, "w") as tb_file:
        rtl_file.write(rtl_output)
        tb_file.write(tb_output)

    return

def main():
    if len(sys.argv) != 5:
        print("Usage: python gen_csr_tb.py <spreadsheet_path> <rtl_output_path> <tb_output_path> <macro_output_path>")
        sys.exit(1)

    spreadsheet_path = sys.argv[1]
    rtl_output_path = sys.argv[2]
    tb_output_path = sys.argv[3]
    macro_output_path = sys.argv[4]

    yaml_file = "./outputs/csr_registers.yml"

    os.makedirs(os.path.dirname(rtl_output_path), exist_ok=True)
    os.makedirs(os.path.dirname(tb_output_path), exist_ok=True)
    os.makedirs(os.path.dirname(macro_output_path), exist_ok=True)
    os.makedirs(os.path.dirname("./outputs"), exist_ok=True)

    excel_to_yaml(spreadsheet_path, yaml_file)
    generate_macros(yaml_file, macro_output_path)
    generate_rtl(yaml_file, rtl_output_path, tb_output_path)

    print(f"[OK] Generated RTL:   {rtl_output_path}")
    print(f"[OK] Generated TB:    {tb_output_path}")
    print(f"[OK] Generated Macros:{macro_output_path}")

if __name__ == "__main__":
    main()
