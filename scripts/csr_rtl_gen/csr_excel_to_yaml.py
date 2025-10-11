#!/usr/bin/env python3
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
import yaml

def excel_to_yaml(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    
    reg_data = {}

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(headers, row))
        name = entry.pop("CSR_NAME").upper()
        reg_data[name] = entry
        reg_data[name]["ADDR_MACRO"] = f"{name.upper()}_ADDR"

    with open(output_file, "w") as f:
        yaml_str = yaml.dump(reg_data, sort_keys=False)
        yaml_str = yaml_str.replace("\n- ", "\n\n- ")

        f.write(yaml_str)

    return

def generate_macros(yaml_file, macro_file):
    with open(yaml_file, "r") as f:
        reg_data = yaml.safe_load(f)

    with open(macro_file, "w") as f:
        f.write("////////////////////////////////////////\n")
        f.write("//         CSR REGISTER MACROS        //\n")
        f.write("////////////////////////////////////////\n")
        
        # figure out longest CSR name for alignment
        max_name_len = max(len(csr_name) for csr_name in reg_data.keys())

        for csr_info in reg_data.values():
            name_field = csr_info['ADDR_MACRO'].ljust(max_name_len + 10)
            addr_field = f"12'h{csr_info['ADDRESS']}"
            line = f"`define {name_field} {addr_field} // {csr_info['DESCRIPTION']}\n"
            f.write(line)

    return

def generate_rtl(yaml_file):
    with open(yaml_file, "r") as f:
        reg_data = yaml.safe_load(f)

    for csr_name, csr_info in reg_data.items():
        reg_data[csr_name]["RESET_VAL"] = f"12'h{csr_info['RESET_VAL']}"
        reg_data[csr_name]["ADDR_MACRO"] = f"`{csr_info['ADDR_MACRO']}"


    max_macro_len = max(len(csr_info["ADDR_MACRO"]) for csr_info in reg_data.values())
    max_name_len = max(len(name) for name in reg_data.keys()) 

    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("csr_regfile_template.sv.j2")
    output = template.render(csr_data=reg_data, max_name_len=max_name_len, max_macro_len=max_macro_len)

    with open("outputs/csr_regfile.sv", "w") as f:
        f.write(output)
    
    return

def main():
    input_file = "/home/wozniak/projects/riscv_pipelined/docs/spreadsheets/csr_spec.xlsx"
    yaml_file = "./outputs/csr_registers.yml"

    excel_to_yaml(input_file, yaml_file)

    generate_macros(yaml_file, "outputs/csr_macros.sv")

    generate_rtl(yaml_file)


if __name__ == "__main__":
    main()
